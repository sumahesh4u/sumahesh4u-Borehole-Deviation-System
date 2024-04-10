from flask import Flask, render_template, jsonify, request, redirect, url_for
from flask_socketio import SocketIO
import hardwareCommunication as hd
import openpyxl
from openpyxl import Workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, Spacer , Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import matplotlib.pyplot as plt
from reportlab.lib import utils
from reportlab.lib.pagesizes import A4
import io
from io import BytesIO 
import os
import time
import pandas as pd
import numpy as np 
from datetime import datetime
from reportlab.lib.utils import ImageReader
from reportlab.platypus import PageBreak
from flaskwebgui import FlaskUI # import FlaskUI
import psutil
from engineio.async_drivers import threading
import logging
#added a comment here
# Configure logging to disable printing HTTP requests and responses
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)


app = Flask(__name__)
socketio = SocketIO(app,async_mode="threading")

# Get the path to the desktop
desktop_path = os.path.expanduser("~/Desktop")

# Ensure the 'uploads' directory exists
uploads_dir = os.path.join(os.getcwd(), 'Backup')
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir)

# Ensure the 'survey's' directory exists
surveys_dir = os.path.join(desktop_path, "Survey's")
if not os.path.exists(surveys_dir):
    os.makedirs(surveys_dir)

#making some global variables:
group_name_global = None
folder_name = None
surveyNameSent = None
counter_value = 0
calibration_interrupted = False
counter_running = False
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def index():
    return render_template('Html/LandingPage.html')

@app.route('/Construction_page')
def under_construction():
    return render_template('Html/Construction_page.html')

@app.route('/button_clicked', methods=['POST'])
def button_clicked():
    button_id = request.form['button_id']

     # Define the URL to Redirect to for each button
    if button_id in ['1']:
        return redirect(url_for(f'page{button_id}'))
    elif button_id in ['2']:
        return redirect(url_for('Data'))
    
    elif button_id in ['3']:
        return redirect(url_for('User_Manual'))

@app.route('/User_Manual')
def User_Manual():
   return render_template('Html/User_Manual.html')
    
@app.route('/get_com_ports')
def get_com_ports():
    com_ports = hd.get_serial_ports()
    return jsonify({'com_port_desc': [port_desc[1] for port_desc in com_ports],'com_port_number': [port_number[0] for port_number in com_ports]})

@app.route('/page1')
def page1():
    return render_template('Html/page1.html')


@app.route('/startSerialCom', methods=['POST'])
def startCommunication():
    jsonData = request.json
    if "body" in jsonData:
        BdsSerialPort = jsonData["body"]
        hd.initializeSerialCom(BdsSerialPort)
        status = hd.checkConnection()[0]
        
        if status == "Your PC is connected to: "+BdsSerialPort:
            
            hd.sendData('l')
            hd.flushSerialBuffers()
        return jsonify({'status': status})
        
    else:
        return jsonify({'success': False, 'message': 'No "body" key in JSON data'})

@app.route('/get_received_messages')
def getReceivedMessages():
    # Assuming you have a function to fetch received messages from hardware
    try:
        
        received_messages = hd.receiveData()
        #hd.flushSerialBuffers()
        
        return jsonify({'messages': received_messages}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
# @app.route('/page2', methods=['GET'])
# def page2():
#     return render_template('Html/page2.html') #if get method is used

@app.route('/page2')
def page2():
    return render_template('Html/page2.html', connection_status="Connected")
@app.route('/page3')
def page3():
    return render_template('Html/page3.html')

# @app.route('/logpage')
# def logpage():
#     return render_template('Html/logpage.html')
@app.route('/submit_survey', methods=['GET','POST'])
def submit_form():
    
    if request.method == 'POST':
        # Get form data
        survey_name = request.form['surveyName']
        hole_id = request.form['holeId']
        start_depth = request.form['startDepth']
        Borehole_Depth = request.form['Borehole Depth']
        Initial_inclination = request.form['Initial inclination']
        Initial_Azimuth = request.form['Initial Azimuth']
        Mineral_Prospect = request.form['Mineral Prospect']
        Northing = request.form['Northing']
        Easting = request.form['Easting']
        Elevation = request.form['Elevation']
        start_interval = request.form['startInterval']
        direction = request.form['direction']
        # Create Excel workbook and add data
        wb = Workbook()
        ws = wb.active
        ws.append(['Label','Entry'])
        ws.append(['Block Name',survey_name])
        ws.append(['Borehole ID',hole_id,])
        ws.append(['Borehole Depth', Borehole_Depth])
        ws.append(['Initial Inclination',  Initial_inclination])
        ws.append(['Initial Azimuth',Initial_Azimuth])
        ws.append(['Mineral Prospect',Mineral_Prospect])
        ws.append(['Northing',Northing])
        ws.append(['Easting',Easting])
        ws.append(['Elevation',Elevation])
        ws.append(['Direction',direction])

        # Save the workbook to a file
        current_datetime = datetime.now()
        formatted_datetime = current_datetime.strftime("%d,%m,%Y_%H,%M,%S")
    
        #adding the folder path to create excel and the name of the excel
        global folder_name
        excel_file_path = os.path.join(surveys_dir+ "\\"+folder_name,  "B" + survey_name + "_" + formatted_datetime + '.xlsx') 
        
        #saving names in global variables
        global survey_name_global
        survey_name_global = hole_id + "_" + formatted_datetime

        wb.save(excel_file_path)

        #sending the name to bds hardware
        global surveyNameSent
        surveyNameSent = "_" + hole_id
        hd.sendData(surveyNameSent)
        #print(hd.sendData(surveyNameSent))
        time.sleep(2) #maintain gap for it to work
        print(hd.ReceiveMultiLineData)
        
        print("survey has started")

    return render_template('Html/startSurveyBDS.html',start_depth=start_depth,start_interval=start_interval,direction=direction)

@app.route('/save_computer_survey_excel', methods=['POST'])
def save_excel():
    try:
        # Retrieve the file data from the request
        file_data = request.files['file'].read()

        # Specify the desired file path
        global folder_name
        global survey_name_global
        file_path = surveys_dir+ "\\"+folder_name+"/computerSurvey.xlsx"  # Update this with your desired path and filename
        
        # Convert the CSV data to a DataFrame
        df = pd.read_csv(io.BytesIO(file_data))

        # Save the DataFrame to an Excel file
        df.to_excel(file_path, index=False)

        return jsonify({'success': True, 'message': 'File saved successfully.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@socketio.on('/batteryLevel')
def getBatteryLevel():
    try:
        print("recived request from front end")
        hd.sendData("b") #requesting bds hardware for battery info
        # Return the battery level in JSON format
        time.sleep(2)
        temp = hd.receiveData()
        print(temp)
        socketio.emit('batteryLevelFromServer',{'batteryLevel':temp})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/Data')
def Data():
    return render_template('Html/Data.html')

@app.route('/Fetch_Data')
def FetchData():
  return render_template('Html/Fetch_Data.html',connection_status="Connected")

@app.route('/group_data', methods=['POST'])
def send_group_name():
        #getting details from the frontend part
        group_name = request.json['group_name']
        location = request.json['location']
        survey_company = request.json['survey_company']
        client = request.json['client']
        client_id = request.json['client_id']
        client_reference = request.json['client_reference']

        #making a unique folder name if the same group does survey again
        global folder_name
        current_datetime = datetime.now()
        formatted_datetime = current_datetime.strftime("%d-%m-%Y-%H-%M-%S")
        folder_name = group_name + "_" + formatted_datetime

        #creating folder based on the given group name by user
        local_path = surveys_dir

        # Join the desktop path and folder name to create the complete path
        folder_path = os.path.join(local_path, folder_name)

        try:
            # Create the folder
            os.makedirs(folder_path)
            print(f"Folder '{folder_name}' created successfully at {folder_path}")
        except FileExistsError:
            print(f"Folder '{folder_name}' already exists at {folder_path}")

        #making the group excel file 
        wb = Workbook()
        ws = wb.active
        ws.append(['Label','Entry'])
        ws.append(['Surveyor Name',group_name])
        ws.append(['Location',location])
        ws.append(['Survey Company',  survey_company])
        ws.append(['Client',client])
        ws.append(['Client ID', client_id])
        ws.append(['Client Reference',client_reference])
        
        # Save the workbook to a file
        Newgroup_Name  = "A"+ group_name
        # current_directory = os.path.abspath(os.path.dirname(__file__))
        excel_file_path = os.path.join(surveys_dir+ "\\"+folder_name, Newgroup_Name + "_" + formatted_datetime + '.xlsx')
        
        #saving names in global variables
        global group_name_global
        group_name_global = group_name  + "_" + formatted_datetime

        wb.save(excel_file_path)

        # render to page3
        return 'ok'
    
@app.route('/startSurveyBDS')
def startsurvey():
  return render_template('Html/startSurveyBDS.html')

@socketio.on('stop_counter')
def stop_counter():
    global counter_running
    global counter_value
    counter_value = 0 #reset the counter
    counter_running = False

    

@socketio.on('stopCalibration')
def stopCalibration():
    global calibration_interrupted
    hd.sendData("z")  # Sending the stop calibration command to the ESP32
    data = hd.receiveData() 
    print("Response from ESP32:", data)
     
    if "Calibration Interrupted" in data:
        print("Calibration stopped")
        socketio.emit('calibrationInterrupted')  # Emit the event if calibration was interrupted
        calibration_interrupted = True
    else:
        print("Unexpected response from ESP32")
        calibration_interrupted = False  # Resetting the flag

@socketio.on('startCalibration')
def startCalibration():
    global calibration_interrupted
    if calibration_interrupted:
        calibration_interrupted = False  # Reset the flag if it was interrupted
    hd.sendData("c")  # Sending the startCalibration command
    # Emitting initial calibration status to start fetching values
    socketio.emit('calibrationOk', {
        'stage1': "0",
        'stage2': "0",
        'stage3': "0",
        'stage4': "0",
    })


@socketio.on('getCalibration')
def getCalibration():
    global calibration_interrupted
    try:
        if not calibration_interrupted :
            # Making some variables here:
            stage1, stage2, stage3, stage4 = None, None, None, None
            data = ""
            
            hd.sendData("c")
            time.sleep(0.01)
            data = hd.receiveData()
            
            
            if "Calibration Interrupted" in data:
                print("Calibration interrupted")
                socketio.emit('calibrationInterrupted')
                calibration_interrupted = True
            else:
                # Parsing the received data to extract calibration stages
                stage1, stage2, stage3, stage4 = map(int, data.split(','))
                print("Calibration status:", stage1, stage2, stage3, stage4)  # For debugging
                
                # Send the calibration status to the webpage
                socketio.emit('calibrationOk', {
                    'stage1': str(stage1),
                    'stage2': str(stage2),
                    'stage3': str(stage3),
                    'stage4': str(stage4),
                })
    except Exception as e:
        print("Error occurred during calibration:", e)
    
# def receive_data_thread():
#   """
#   This function retrieves data from the serial port using hd.receiveData()
#   and prints it. It runs in a separate thread.
#   """
#   while True:
#     x = hd.receiveData()
#     print(x)
        
@socketio.on('start_counter')
def start_counter():
    print("Received request from front end")
    global counter_value
    global counter_running

    # Send the command to BDS hardware to start the counter
    hd.sendData("s")
    print("Command sent")
    time.sleep(0.1)
    # Wait for a response from BDS hardware
    data = hd.receiveData()  # Adjust the timeout value as needed

    if data == "y":
        print("Acknowledgement received: ", data)
        counter_running = True
        # # Thread creation and start
        # receive_thread = threading.Thread(target=receive_data_thread)
        # receive_thread.start()

        start_time = time.time()  # Start the timer
        
        while counter_running:
            counter_value = time.time() - start_time
            socketio.emit('update_counter', {'value': round(counter_value*1000)})
            #counter_value += 1
            socketio.sleep(0.0000001)  # Adjust the timer as needed
    else:
        print("Wrong acknowledgement received: ", data)




  
@app.route('/return-to-page1')
def return_to_page1():
    return redirect(url_for('index'))
    
@socketio.on('/Read_data')
def Read_data():
    global folder_name
    global surveyNameSent

    try:
        if surveyNameSent != None:
            file_path = surveys_dir+ "\\"+folder_name+"/HardwareSurveyData.xlsx"

            #making the group excel file 
            wb = Workbook()
            ws = wb.active

            ws.append(["TimeStamp","Pitch","ROll","Yaw"]) #adding the headings

            # hd.sendData("i") #stop timer
            # print(hd.receiveData())
            # time.sleep(2)

            temp = list(surveyNameSent)  # Convert string to a list of characters
            temp[0] = "/"  # Replace the first character ('_') with '/'
            temp = ''.join(temp)  # Convert the list back to a string
            temp += ".csv"  # Append ".csv" to the string
            hd.sendData("esc")
            time.sleep(0.01)
            print("file name should be: ",temp)
            hd.sendData("r "+temp) #send the name like: r /fileName.csv
            
            line_count = 0
            socketio.emit('/readingStarted',{"started" : "yes"})

            while True:
                temp = hd.receiveData()
                print("Received:", temp)  # Debugging statement

                line_count += 1# Increment the line count
                
                if line_count <= 2:# Ignore the first two lines
                    continue

                if temp != "dataEnded":
                    tempList = temp.split(',')
                    ws.append(tempList) #append the list now
                else:
                    print("End of data received.")  # Debugging statement
                    break
            
            # Save the workbook to a file
            wb.save(file_path)

            # Emit the file path to the client if the file is successfully written
            socketio.emit('/data_received', {'file_path':file_path})
        else:
            print("file name is not given so assumed directly fetching the data",surveyNameSent)
            socketio.emit('/data_received', {'file_path':"none"})
    except Exception as e:
        # Handle file writing errors
        print(f"Error writing file: {e}")
        # Emit an error event to the client for error handling
        socketio.emit('file_error', str(e))
        


 
@socketio.on('Readfile')
def handle_read_file(data):
    file_name = data.get('file_name')
    folder_path = data.get('folder_path')
    folder_path1 =  os.path.join(surveys_dir+ "\\"+folder_path)
        
    print("file_name")
    print(folder_path1)
    print(file_name)
    # Ensure the directory exists
    if not os.path.exists(folder_path1):
        raise ValueError("Folder path does not exist.")
    
    # Constructing the file path using folder_path received from the client
    file_path = os.path.join(folder_path1, "HardwareSurveyData.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["TimeStamp", "Pitch", "ROll", "Yaw"]) # Add the headings
    
    # Assuming 'hd' is defined somewhere in your code
    hd.sendData("esc")
    time.sleep(0.01)
    hd.sendData(f"r /{file_name}.csv") # Send the name like: r /fileName.csv
   
    line_count = 0
    socketio.emit('/readingStarted', {"started": "yes"})

    while True:
        temp = hd.receiveData()
        print("Received:", temp)  # Debugging statement

        line_count += 1 # Increment the line count
            
        if line_count <= 2: # Ignore the first two lines
            continue

        if temp != "dataEnded":
            tempList = temp.split(',')
            ws.append(tempList) # append the list now
        else:
            print("End of data received.")  # Debugging statement
            break
        
    # Save the workbook to a file
    wb.save(file_path)

    # Emit the file path to the client if the file is successfully written
    socketio.emit('/data_received', {'file_path': file_path})

    # # Save the workbook to a file
    # ws.save(file_path)

    # Emit the file path to the client if the file is successfully written
    socketio.emit('/data_received', {'file_path': file_path})




@socketio.on('clear_data')
def clear_data():
    hd.sendData("d")
    time.sleep(2)
    message = hd.receiveData()
    print(message)
    socketio.emit('data_sent', {'message':message})
    #print(message)

@app.route('/terminalCommand',methods=["POST"])
def terminalCommands():
    print("command received from server:",request.json["cmd"])
    hd.sendData(request.json["cmd"]) #send whatever command is sent by client to bds hardware
    time.sleep(2) #give some gap for it work
    temp = hd.ReceiveMultiLineData()
    print(temp)
    return temp
# Define a route to stream live serial data
@app.route('/stream_serial_data')
def stream_serial_data():
    # Read serial data in a loop and yield it as a response
    while True:
        data=hd.receiveData()
        yield f"data: {data}\n\n"

#######################################################################
######################################################################
def copy_directory(source, destination):
    try:
        # Check if the source directory exists
        if os.path.exists(source):
            # Extracting the directory name from the source path
            directory_name = os.path.basename(source)
            # Constructing the destination path with the directory name
            destination_with_dirname = os.path.join(destination, directory_name)
            
            # Constructing the shell command to copy the directory and its contents
            command = f"xcopy /E /I /Y \"{source}\" \"{destination_with_dirname}\""
            # Using os.system to execute the command
            os.system(command)
            
            print(f"Directory '{source}' copied to '{destination_with_dirname}' successfully.")
        else:
            print(f"Source directory '{source}' does not exist.")
    except Exception as e:
        print(f"Error: {e}")

def find_processes_using_file(file_path):
    processes = []
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            open_files = proc.open_files()
            for of in open_files:
                if of.path == file_path:
                    processes.append((proc.pid, proc.name()))
        except psutil.NoSuchProcess:
            pass
    return processes

def remove_directory(directory):
    try:
        # Constructing the shell command to remove the directory and its contents
        command = f"rmdir /s /q \"{directory}\""
        # Using os.system to execute the command
        os.system(command)
        print(f"Directory '{directory}' and its contents removed successfully.")
    except Exception as e:
        print(f"Error: {e}")

################################################################################
################################################################################
 
@app.route('/run_python_code', methods=['POST'])
def run_python_code():
    files = request.files.getlist('file')  # Use getlist to get all files

    if len(files) == 4:
        # Create a list to store file paths
        file_paths = []
        for file in files:
            # Save the file to the 'uploads' directory
            file_path = os.path.join(uploads_dir, file.filename)
            file.save(file_path)
            file_paths.append(file_path)
        # Now you have the actual file paths on the server side
        print("Received file paths:", file_paths)

        file1_path = file_paths[-2]
        file2_path = file_paths[-1]
        file3_path = file_paths[0]
        file4_path = file_paths[1]

        result_message = f"Report Sucessfully Generated on Desktop"

        df = pd.read_excel(file2_path)
        ranges_df = pd.read_excel(file1_path) 

        ranges_list = ranges_df['TimeStamp'].unique()  
        ranges = [(value, value + 30000) for value in ranges_list]
        result_data = []

        for start_range, end_range in ranges:
            filtered_df = df[(df['TimeStamp'] >= start_range+10000) & (df['TimeStamp'] <= end_range)]
            
             # Convert angles to radians for circular mean calculation
            pitch_radians = np.radians(filtered_df['Pitch'])
            roll_radians = np.radians(filtered_df['ROll'])
            yaw_radians = np.radians(filtered_df['Yaw'])

            # Circular mean calculation for Pitch, Roll, and Yaw
            avg_column2_rad = np.arctan2(np.mean(np.sin(pitch_radians)), np.mean(np.cos(pitch_radians)))
            avg_column3_rad = np.arctan2(np.mean(np.sin(roll_radians)), np.mean(np.cos(roll_radians)))
            avg_column4_rad = np.arctan2(np.mean(np.sin(yaw_radians)), np.mean(np.cos(yaw_radians)))

         # Convert back to degrees
            avg_column2 = np.degrees(avg_column2_rad)
            avg_column3 = np.degrees(avg_column3_rad)
            avg_column4_deg = np.degrees(avg_column4_rad)

            # Ensure the angles are within [0, 360) range
            # avg_column2 = (avg_column2_deg + 360) % 360
            # avg_column3 = (avg_column3_deg + 360) % 360
            avg_column4 = (avg_column4_deg + 360) % 360

            # Calculate inclination using circular statistics
            inclination = avg_column2#np.sqrt(avg_column2**2 + avg_column3**2)

            result_data.append({
                'TimeStamp': f'{start_range}',
                'MarkStation': ranges_df[ranges_df['TimeStamp'] == start_range]['MarkStation'].values[0],
                'Pitch': avg_column2,
                'Roll': avg_column3,
                'Yaw': avg_column4,
                'Inclination': inclination
            })

        result_df = pd.DataFrame(result_data)
        result_df.to_excel(file1_path, index=False)

        excel_file_path = file3_path
        table_excel_file_path = file1_path
        additional_excel_file_path = file4_path  # Add Book3 file path

        # Extract the name of the first file
        original_first_file_name = os.path.splitext(os.path.basename(file_paths[0]))[0]

        # Remove the first character from the extracted name 
        modified_first_file_name = original_first_file_name[1:]
        modified_first_file_name = modified_first_file_name[:-19]
        #print(modified_first_file_name)

        # get the bore hold id
        # Read the value from the second column, second row of the second file
        second_file = pd.read_excel(file4_path)
        second_value = second_file.iloc[1, 1]  # Assuming indexing starts from 0

        # Concatenate the value with the survey report PDF
        pdf_output_path = os.path.join(desktop_path, f"{modified_first_file_name}_{second_value}_Survey_Report.pdf")

        # Set the PDF output path with the modified name on the desktop
        #pdf_output_path = os.path.join(desktop_path, f"{modified_first_file_name}_Survey_Report.pdf")


        # Read data from Excel
        data = read_excel(excel_file_path)

        # Read table data from another Excel file
        table_data = read_table_excel(table_excel_file_path)

        # Read additional data from another Excel file
        additional_data = read_additional_excel(additional_excel_file_path)

        # Generate PDF report
        generate_pdf(data, table_data, additional_data, pdf_output_path)

        file_paths = [file1_path,file2_path,file3_path,file4_path]
        output_file = os.path.join(uploads_dir, f"{modified_first_file_name}_merged_file.xlsx") 
        merge_excel_files(file_paths, output_file)

        # # Clean up temporary files
        # os.remove(file1_path)
        # os.remove(file2_path)
        # os.remove(file3_path)
        # os.remove(file4_path)

        survey_folder = surveys_dir
        try:
            for filename in os.listdir(survey_folder):
                src = os.path.join(survey_folder, filename)
                print("src folder:",src)
            
            copy_directory(src, uploads_dir)

        except Exception as err:
            print(err)

        return jsonify({'message': result_message})
    else:
        return jsonify({'error': 'Please upload exactly four files.'})

@app.route("/delete_survey_folder",methods=['POST'])
def delete_survey_folder():
    print("req to delete recevied!!")
    remove_directory(surveys_dir)
    return jsonify({'message': "file deleted successfully"})

def merge_excel_files(file_paths, output_file):
    # Create a Pandas Excel writer using openpyxl
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Iterate through each Excel file and read data into a Pandas DataFrame
        for file_path in file_paths:
            sheet_name = f"Sheet_{file_paths.index(file_path) + 1}"  # Sheet names can be customized
            df = pd.read_excel(file_path)
            
            # Write the DataFrame to the Excel file as a new sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, value = row
        data.append((name, value))
    return data

def read_table_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    table_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        col1, col2, col3, col4, col5, col6 = row
        table_data.append((col1, col2, col3, col4, col5, col6))
    return table_data

def read_additional_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    additional_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, value1 = row
        additional_data.append((name, value1))
    return additional_data

def add_watermark(canvas, watermark_path, opacity=0.5):
    """
    Add a watermark to every page of the PDF.
    """
    canvas.saveState()
    watermark = ImageReader(os.path.join(os.getcwd(), 'watermark.png'))
    iw, ih = watermark.getSize()
    c_width, c_height = canvas._pagesize
    scale = 0.25*min(c_width/iw, c_height/ih)
    w = iw * scale
    h = ih * scale
    canvas.setFillColorRGB(1, 1, 1, opacity)
    canvas.drawImage(watermark, (c_width - w) / 2, (c_height - h) / 2, width=w, height=h)
    canvas.restoreState()

def generate_pdf(data, table_data, additional_data, pdf_path):
    doc = SimpleDocTemplate(pdf_path, pagesize=A4)
    # Add watermark function to each page
    watermark_path = os.path.join(os.getcwd(), 'watermark.png')
    watermark_opacity=0.1
    doc.build(
        [Spacer(1, 1)],
        onFirstPage=lambda canvas, _: add_watermark(canvas, watermark_path, watermark_opacity),
        onLaterPages=lambda canvas, _: add_watermark(canvas, watermark_path, watermark_opacity),
    )
    elements = []

    # Setup styles
    styles = getSampleStyleSheet()

    # Define styles for heading and normal text
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName='Helvetica-Bold', alignment=1, fontSize=20)
    heading_style = ParagraphStyle('Heading1', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=12)
    sig_style = ParagraphStyle('Heading1', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14, alignment=2)  
    underline_heading_style = ParagraphStyle('Heading1Underline', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14, underline=1)

##########################################################################################################################
##########################################################################################################################
    # Add an image to the report
    image_path = "static/Assets/images/header1.png"  # Replace with the actual path to your image file

    # Get the width and height of the image while maintaining the aspect ratio
    img_width, img_height = utils.ImageReader(image_path).getSize()

    # Calculate the aspect ratio
    aspect_ratio = img_width / img_height

    # Set the width and height based on the aspect ratio
    max_width = doc.width   # Adjust the maximum width based on your layout
    max_height = doc.width  # Adjust the maximum height based on your layout

    if aspect_ratio > 1:
        img = Image(image_path, width=max_width, height=max_width / aspect_ratio, hAlign='RIGHT')
    else:
        img = Image(image_path, width=max_height * aspect_ratio, height=max_height, hAlign='RIGHT')

    elements.append(img)

    # Add separator line
    line_width = doc.width
    elements.append(Paragraph("_" * int(line_width/7), normal_style))
    elements.append(Spacer(0,8))  # Adjust the line sapcing
#########################################################################   #################################################
############### ###########################################################################################################
    # Add centered title at the start of the report
    title = "Deviation Survey Report"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(0,8))  # Adjust the line sapcing
    # Add separator line
    #elements.append(Paragraph("_" * int(line_width/7), normal_style))
    #elements.append(Spacer(0,8))  # Adjust the line sapcing
##########################################################################################################################
##########################################################################################################################
    # Add data to the report
    elements.append(Paragraph("Group Details:", heading_style))
    #for name, value in data:
        # Create data_table
    data_table = [[Paragraph(f"{name}", normal_style), Paragraph(":&nbsp;"f"{value}", normal_style)] for name, value in data]

        # Add data to the report using a table
    table_style = [('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]
    table = Table(data_table, style=table_style)
    elements.append(table)
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(Paragraph("_" * int(line_width/7), normal_style))
    elements.append(Spacer(0,8))  # Adjust the line sapcing       
##########################################################################################################################
##########################################################################################################################
    # Add data to the report
    elements.append(Paragraph("Survery Details:", heading_style))
    #for name, value in data:
        # Create data_table
    survey_data_table = [[Paragraph(f"{name}", normal_style), Paragraph(":&nbsp;"f"{value}", normal_style)] for name, value in additional_data]
    elements.append(Table(survey_data_table, style=table_style))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(Paragraph("_" * int(line_width/7), normal_style))
    elements.append(Spacer(0,8))  # Adjust the line sapcing 
##########################################################################################################################
##########################################################################################################################
    # Extract only the 2nd, 5th, and 6th columns
    filtered_table_data = [[row[1], row[5], row[4]] for row in table_data]

    # Add a new column for serial numbers
    filtered_table_data.insert(0, ['S.No', 'Depth in Meters','Inclination Angle', 'Azimuth Angle'])
    for i, row in enumerate(filtered_table_data[1:], start=1):
        row.insert(0, i)
    for row in filtered_table_data:
        for i, value in enumerate(row):
            if isinstance(value, float):
                row[i] = round(value, 2)
    # Create the table
    # Create the table
    table_style = [
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Highlight first row
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')   # Bold first row
    ]

    table = Table(filtered_table_data, style=table_style)
    elements.append(PageBreak())
    # Add the modified table to the report
    elements.append(Paragraph("Borehole Data:", heading_style))
    elements.append(table)
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(Paragraph("_" * int(line_width/7), normal_style))
    elements.append(Spacer(0, 8))  # Adjust the line spacing   
##########################################################################################################################
##########################################################################################################################
    elements.append(PageBreak())
      # Add new section "Plots"
    elements.append(Paragraph("Plots:", heading_style))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
##########################################################################################################################
##########################################################################################################################
    # Create plot between first and second columns
    plt.figure(figsize=(6, 4))
    plt.plot([row[1] for row in table_data], [row[5] for row in table_data], label='Inclination')
    plt.xlabel('Depth in Meters')
    plt.ylabel('Inclination Angle')
    plt.legend()
    plt.title('Plot between Inclination Angle and Depth in Meters')

    # Save plot to BytesIO buffer
    buffer1 = BytesIO()
    plt.savefig(buffer1, format='png')
    buffer1.seek(0)
    plt.close()

    # Add plot to the report
    elements.append(Paragraph("<u>Plot between Inclination Angle and Depth in Meters</u>:", underline_heading_style))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(Image(buffer1, width=300, height=200, hAlign='CENTER'))
    #elements.append(Image(buffer1, width=doc.width-20, height=doc.width-20, hAlign='CENTER'))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(PageBreak())
##########################################################################################################################
##########################################################################################################################
    # Create plot between first and third columns
    plt.figure(figsize=(6, 4))
    plt.plot([row[1] for row in table_data], [row[4] for row in table_data], label='Azimuth')
    plt.xlabel('Depth in Meters')
    plt.ylabel('Azimuth Angle')
    plt.legend()
    plt.title('Plot between Azimuth Angle and Depth in Meters')

    # Save plot to BytesIO buffer
    buffer2 = BytesIO()
    plt.savefig(buffer2, format='png')
    buffer2.seek(0)
    plt.close()

    # Add plot to the report
    elements.append(Paragraph("<u>Plot between Azimuth Angle and Depth in Meters</u>:", underline_heading_style))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(Image(buffer2, width=300, height=200, hAlign='CENTER'))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(PageBreak())
##########################################################################################################################
##########################################################################################################################
    # Create a 3D plot between first, fourth, and fifth columns
    fig = plt.figure(figsize=(8, 6))
    ax = fig.add_subplot(111, projection='3d')
    #ax.scatter([-row[5] for row in table_data], [row[4] for row in table_data], [-row[1] for row in table_data], label='Column 3D')  # Adjusted data for inverted z-axis
    ax.plot([row[5] for row in table_data], [row[4] for row in table_data], [-row[1] for row in table_data], marker='o', linestyle='-', label='3D plot')  # Adjusted data for inverted z-axis

    # Set labels for each axis
    ax.set_xlabel('Inclination Angle')
    ax.set_ylabel('Azimuth Angle')  # Add your appropriate label for the y-axis
    ax.set_zlabel('Depth in Meters')

    ax.legend()
    ax.set_title('3D Plot between Azimuth Angle, Depth in Meters, and Inclination Angle')

   # Customize the orientation of the axes
    ax.view_init(10,45,0,"z") # Axes3D.view_init(elev=None, azim=None, roll=None, vertical_axis='z', share=False)
                                #https://matplotlib.org/stable/api/_as_gen/mpl_toolkits.mplot3d.axes3d.Axes3D.view_init.html
    # Save the 3D plot to BytesIO buffer
    buffer3 = BytesIO()
    plt.savefig(buffer3, format='png')
    buffer3.seek(0)
    plt.close()

    # Add 3D plot to the report
    elements.append(Paragraph("<u>3D Plot between Azimuth Angle, Depth in Meters, and Inclination Angle</u>:", underline_heading_style))
    elements.append(Spacer(0, 8))  # Adjust the line spacing
    elements.append(Image(buffer3, width=400, height=300, hAlign='CENTER'))
    elements.append(Spacer(0, 8))
    elements.append(PageBreak())
##########################################################################################################################
##########################################################################################################################
    elements.append(Paragraph("<u>Bulls Eye Representation</u>:", underline_heading_style))
    # Create polar plot
    depths = [row[1] for row in table_data]
    inclinations = [row[5] for row in table_data]
    azimuths = [row[4] for row in table_data]
    inclinations_rad = np.radians(inclinations)

    fig, ax = plt.subplots(subplot_kw={'projection': 'polar'})
    scatter = ax.scatter(inclinations_rad, depths, c=azimuths, cmap='viridis', s=100, alpha=0.75)
    ax.set_theta_zero_location("N")
    ax.set_theta_direction(-1)

    # Add colorbar
    cbar = plt.colorbar(scatter, ax=ax, orientation='vertical', pad=0.1)
    cbar.set_label('Azimuth Angle')

    # Save plot to BytesIO buffer
    buffer4 = BytesIO()
    plt.savefig(buffer4, format='png')
    buffer4.seek(0)
    plt.close()

    # Add plot to the report
    elements.append(Image(buffer4, width=doc.width-20, height=doc.width-20, hAlign='CENTER'))

##########################################################################################################################
##########################################################################################################################
    elements.append(Spacer(0, 200))  # Increase the vertical space
    elements.append(Paragraph("Signature of the concerned authority", sig_style))

    # Build the PDF report
    doc.build(elements)
 
if __name__ == '__main__':
    
    FlaskUI(
        app=app,
        socketio=socketio,  # Assuming you're using websockets
        server="flask_socketio",
        width=1024,
        height=768,
    ).run()
    